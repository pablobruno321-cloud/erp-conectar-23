"""
Exportación Excel — backup legible por tablas (Backup 2).
"""

from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

from .models import (
    Cliente, Proveedor, ProveedorContacto, ProveedorLogistico,
    Producto, Cotizacion, Pedido, ItemPedido,
    Cobranza, CobranzaLinea, CobranzaImputacion,
    FormaCobro, Fruta, Variedad, Clasificacion, Envase, Marca, Kilogramo,
)

HEADER_FILL = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _fmt_dt(val):
    if not val:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y %H:%M')
    if isinstance(val, date):
        return val.strftime('%d/%m/%Y')
    return str(val)


def _si_no(val):
    return 'Sí' if val else 'No'


def _hoja(wb, titulo, headers, filas):
    ws = wb.create_sheet(titulo)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for fila in filas:
        ws.append(fila)
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 50)
    return ws


def crear_backup_excel(dest_dir: Path | str | None = None):
    """
    Crea backup Excel con una hoja por tabla principal.
    Returns: (True, ruta) o (False, mensaje_error)
    """
    if dest_dir is None:
        dest_dir = Path(__file__).resolve().parent.parent / 'backups' / 'excel'
    else:
        dest_dir = Path(dest_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    archivo = dest_dir / f'backup_{timestamp}.xlsx'

    wb = Workbook()
    wb.remove(wb.active)

    _hoja(wb, 'Clientes',
          ['ID', 'Nombre', 'Mercado', 'Puesto', 'Contacto', 'CUIT', 'Teléfono', 'Email', 'Dirección', 'Saldo', 'Activo', 'Fecha Creación'],
          [[c.id, c.nombre, c.mercado, c.puesto, c.contacto, c.cuit, c.telefono, c.email, c.direccion,
            c.saldo, _si_no(c.activo), _fmt_dt(c.fecha_creacion)] for c in Cliente.query.order_by(Cliente.id)])

    _hoja(wb, 'Proveedores',
          ['ID', 'Nombre', 'Razón Social', 'CUIT', 'Teléfono', 'Email', 'Dirección', 'Provincia', 'Saldo', 'Activo', 'Fecha Creación'],
          [[p.id, p.nombre, p.razon_social, p.cuit, p.telefono, p.email, p.direccion, p.provincia,
            p.saldo, _si_no(p.activo), _fmt_dt(p.fecha_creacion)] for p in Proveedor.query.order_by(Proveedor.id)])

    _hoja(wb, 'Prov_Contactos',
          ['ID', 'Proveedor ID', 'Proveedor', 'Nombre', 'Rol', 'Teléfono', 'Email', 'Contacto Logístico'],
          [[c.id, c.proveedor_id, c.proveedor.nombre if c.proveedor else '', c.nombre, c.rol, c.telefono, c.email,
            _si_no(c.contacto_logistico)] for c in ProveedorContacto.query.order_by(ProveedorContacto.id)])

    _hoja(wb, 'Prov_Logisticos',
          ['ID', 'Nombre', 'CUIT', 'Teléfono', 'Email', 'Tarifa', 'Activo', 'Fecha Creación'],
          [[p.id, p.nombre, p.cuit, p.telefono, p.email, p.tarifa, _si_no(p.activo), _fmt_dt(p.fecha_creacion)]
           for p in ProveedorLogistico.query.order_by(ProveedorLogistico.id)])

    _hoja(wb, 'Productos',
          ['ID', 'Proveedor ID', 'Proveedor', 'Fruta', 'Variedad', 'Clasificación', 'Envase', 'Kg', 'Marca', 'Costo Unit.', 'Activo'],
          [[pr.id, pr.proveedor_id, pr.proveedor.nombre if pr.proveedor else '', pr.fruta, pr.variedad,
            pr.clasificacion, pr.envase, pr.kilogramo, pr.marca, pr.costo_unitario, _si_no(pr.activo)]
           for pr in Producto.query.order_by(Producto.id)])

    _hoja(wb, 'Cotizaciones',
          ['ID', 'Producto ID', 'Producto', 'Desde', 'Hasta', 'Costo', 'Precio', 'Margen %', 'Activo'],
          [[c.id, c.producto_id, c.producto.fruta if c.producto else '', _fmt_dt(c.fecha_desde), _fmt_dt(c.fecha_hasta),
            c.costo_unitario, c.precio_unitario, c.margen_ganancia, _si_no(c.activo)]
           for c in Cotizacion.query.order_by(Cotizacion.id)])

    _hoja(wb, 'Pedidos',
          ['ID', 'Número', 'Cliente ID', 'Cliente', 'Fecha Venta', 'Fecha Carga', 'Remito', 'Fecha Remito',
           'Entregado', 'Fecha Entrega', 'Eliminado', 'Costo Total', 'Venta Total', 'Comisión', 'Resultado',
           'Desc. Costo', 'Desc. Comisión', 'Prov. Logístico'],
          [[p.id, p.numero, p.cliente_id, p.cliente.nombre if p.cliente else '', _fmt_dt(p.fecha_venta),
            _fmt_dt(p.fecha_carga), p.remito, _fmt_dt(p.fecha_remito), _si_no(p.entregado), _fmt_dt(p.fecha_entrega),
            _si_no(p.eliminado), p.costo_total, p.precio_venta_total, p.comision, p.resultado,
            p.descuento_costo, p.descuento_comision,
            p.prov_logistico.nombre if p.prov_logistico else ''] for p in Pedido.query.order_by(Pedido.id)])

    _hoja(wb, 'Items_Pedido',
          ['ID', 'Pedido ID', 'N° Pedido', 'Producto ID', 'Proveedor', 'Fruta', 'Pallets', 'Bultos', 'Kg',
           'Cantidad', 'Unidad', 'Costo Unit.', 'Precio Unit.', 'Costo Total', 'Precio Total', 'Comisión', 'Resultado'],
          [[it.id, it.pedido_id, it.pedido.numero if it.pedido else '', it.producto_id,
            it.producto.proveedor.nombre if it.producto and it.producto.proveedor else '',
            it.producto.fruta if it.producto else '', it.pallets, it.bultos, it.kg, it.cantidad, it.unidad,
            it.costo_unitario, it.precio_unitario, it.costo_total, it.precio_total, it.comision, it.resultado]
           for it in ItemPedido.query.order_by(ItemPedido.id)])

    _hoja(wb, 'Cobranzas',
          ['ID', 'Cliente ID', 'Cliente', 'Fecha', 'Monto', 'Método', 'Notas', 'Eliminado', 'Fecha Eliminación'],
          [[c.id, c.cliente_id, c.cliente.nombre if c.cliente else '', _fmt_dt(c.fecha_cobranza), c.monto,
            c.metodo, c.notas, _si_no(c.eliminado), _fmt_dt(c.fecha_eliminacion)] for c in Cobranza.query.order_by(Cobranza.id)])

    _hoja(wb, 'Cobranzas_Lineas',
          ['ID', 'Cobranza ID', 'Forma', 'Monto', 'Concepto'],
          [[ln.id, ln.cobranza_id, ln.forma, ln.monto, ln.concepto] for ln in CobranzaLinea.query.order_by(CobranzaLinea.id)])

    _hoja(wb, 'Cobranzas_Imput',
          ['ID', 'Cobranza ID', 'Pedido ID', 'N° Pedido', 'Monto Imputado', 'Pago Proveedor', 'Comisión', 'Fecha'],
          [[i.id, i.cobranza_id, i.pedido_id, i.pedido.numero if i.pedido else '', i.monto_imputado,
            i.pago_proveedor, i.comision, _fmt_dt(i.fecha_conciliacion)] for i in CobranzaImputacion.query.order_by(CobranzaImputacion.id)])

    _hoja(wb, 'Formas_Cobro',
          ['ID', 'Nombre', 'Activo'],
          [[f.id, f.nombre, _si_no(f.activo)] for f in FormaCobro.query.order_by(FormaCobro.id)])

    for modelo, nombre in (
        (Fruta, 'Frutas'), (Variedad, 'Variedades'), (Clasificacion, 'Clasificaciones'),
        (Envase, 'Envases'), (Marca, 'Marcas'), (Kilogramo, 'Kilogramos'),
    ):
        _hoja(wb, nombre, ['ID', 'Nombre', 'Activo'],
              [[m.id, m.nombre, _si_no(m.activo)] for m in modelo.query.order_by(modelo.id)])

    try:
        wb.save(archivo)
        return True, str(archivo)
    except Exception as e:
        return False, str(e)
